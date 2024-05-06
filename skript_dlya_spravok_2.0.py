import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import psycopg2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime

def generate_excel():
    # Получение параметров из интерфейса
    company_name = company_name_var.get()
    start_date = start_date_var.get()
    end_date = end_date_var.get()

    # Генерация SQL-запроса на основе параметров
    sql_query = f"""
    SELECT d.id as did, e.personnel_number as epn,
    e.surname as es, e."name" as en, e.patronymic as ep,
    o.id as "oid", o."name" as "on", d.resolution as dr,
    d.diagnosis as dd, c."content" as cc, d.created as dc
    FROM structures.employees e
    INNER JOIN structures.organizations o ON e.org_id = o.id
    INNER JOIN medrec.documents d ON d.employee_id = e.id
    INNER JOIN medrec."comments" c ON c.reference_id = d.id
    WHERE o."name" like '%{company_name}%'
    AND d.created BETWEEN '{start_date} 00:00:00 +0300' AND '{end_date} 23:59:59.999 +0300'
    AND c.reference_type = 'document';
    """

    # Создание подключения к базе данных и выполнение запроса
    try:
        conn = psycopg2.connect(**db_params)
        cur = conn.cursor()
        cur.execute(sql_query)
        results = cur.fetchall()
        cur.close()
        conn.close()

        # Создание нового Excel-файла и запись данных
        workbook = Workbook()
        sheet = workbook.active
        for row_index, row_data in enumerate(results, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.astimezone().replace(tzinfo=None)
                sheet.cell(row=row_index, column=col_index, value=cell_value)

        workbook.save("output.xlsx")
        messagebox.showinfo("Успех", "Данные успешно экспортированы в Excel.")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

# Параметры подключения к базе данных PostgreSQL
db_params = {
    'host': '****',
    'database': '****',
    'user': '****',
    'password': '****'
}

# Создание графического интерфейса
app = tk.Tk()
app.title("Экспорт данных в Excel")

company_name_label = ttk.Label(app, text="Название компании:")
company_name_label.pack()
company_name_var = tk.StringVar()
company_name_entry = ttk.Entry(app, textvariable=company_name_var)
company_name_entry.pack()

start_date_label = ttk.Label(app, text="Начальная дата (гггг-мм-дд):")
start_date_label.pack()
start_date_var = tk.StringVar()
start_date_entry = ttk.Entry(app, textvariable=start_date_var)
start_date_entry.pack()

end_date_label = ttk.Label(app, text="Конечная дата (гггг-мм-дд):")
end_date_label.pack()
end_date_var = tk.StringVar()
end_date_entry = ttk.Entry(app, textvariable=end_date_var)
end_date_entry.pack()

generate_button = ttk.Button(app, text="Сгенерировать Excel", command=generate_excel)
generate_button.pack()

app.mainloop()
